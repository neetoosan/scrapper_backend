"""
XLSX export module — generates Excel workbooks from scrape results.
Port of excel-builder.js using openpyxl.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Header styling
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style='thin', color='1A3A6B'),
)

# Column definitions: (header, field_name, width)
COLUMNS = [
    ('Name', 'name', 28),
    ('Company', 'company_name', 28),
    ('Phone', 'phone_numbers', 24),
    ('WhatsApp', 'whatsapp_numbers', 24),
    ('Social Handle', 'social_media_handles', 32),
    ('Email', 'emails', 32),
    ('Website', 'website', 36),
    ('Address', 'address', 40),
    ('Category', 'category', 20),
    ('Rating', 'rating', 10),
    ('Reviews', 'review_count', 12),
    ('Source URL', 'source_url', 40),
    ('Page Title', 'page_title', 28),
]


def build_workbook(listings: list[dict], page_title: str = '', page_url: str = '') -> bytes:
    """
    Build an XLSX workbook from a list of contact/business listings.
    Returns the workbook as bytes for streaming download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    # Info rows
    ws.append(['Page Title', page_title or 'Scraped Contacts'])
    ws.append(['Page URL', page_url or ''])
    ws.append([])  # Empty row

    # Header row
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)

    # Style header row (row 4)
    header_row = 4
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    for listing in listings:
        row = []
        for _, field, _ in COLUMNS:
            value = listing.get(field, '')
            # Join lists into comma-separated strings
            if isinstance(value, list):
                value = ', '.join(str(v) for v in value if v)
            row.append(str(value) if value else '')
        ws.append(row)

    # Auto-filter on header row
    if listings:
        last_col_letter = ws.cell(row=header_row, column=len(COLUMNS)).column_letter
        ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{header_row + len(listings)}'

    # Freeze panes below header
    ws.freeze_panes = f'A{header_row + 1}'

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
